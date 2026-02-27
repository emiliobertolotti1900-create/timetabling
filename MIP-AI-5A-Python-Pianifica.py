import math
import re
from collections import defaultdict

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

from scipy.optimize import milp, LinearConstraint, Bounds
import scipy.sparse as sp


INPUT_XLSX = "MIP-AI-3B.xlsx"
OUTPUT_XLSX = "Piano_Risolto.xlsx"


def parse_list_cell(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return []
    s = str(x).strip()
    if not s:
        return []
    parts = re.split(r"[;,]+", s)
    return [p.strip() for p in parts if p.strip()]


def parse_int_list(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return []
    s = str(x)
    parts = re.split(r"[;,]+", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # gestisce "2, 4" e simili
        for q in re.split(r"\s+", p):
            q = q.strip()
            if q.isdigit():
                out.append(int(q))
    # unique preservando ordine
    seen = set()
    res = []
    for v in out:
        if v not in seen:
            seen.add(v)
            res.append(v)
    return res


def parse_id_list(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return []
    s = str(x).strip()
    if not s:
        return []
    parts = re.split(r"[;,]+", s)
    return [p.strip() for p in parts if p.strip()]


def parse_matrix_presence(ws, id_col=1, start_row=3):
    """
    Ritorna:
      pres[entity_id][(day, bucket_descr)] = True se PRESENTE, False se ASSENTE.
    Struttura foglio: riga1 giorni (merge), riga2 bucket, dati da riga3.
    """
    col_map = {}
    current_day = None
    for c in range(3, ws.max_column + 1):
        d = ws.cell(1, c).value
        if d is not None:
            current_day = str(d).strip()
        b = ws.cell(2, c).value
        if b is None or current_day is None:
            continue
        col_map[c] = (current_day, str(b).strip())

    pres = defaultdict(dict)
    r = start_row
    while True:
        eid = ws.cell(r, id_col).value
        if eid is None:
            break
        eid = str(eid).strip()
        for c, (day, buck) in col_map.items():
            val = ws.cell(r, c).value
            present = (val is None) or (isinstance(val, float) and math.isnan(val)) or (str(val).strip() == "")
            pres[eid][(day, buck)] = present
        r += 1
    return pres


def main():
    # --- Caricamento workbook (per layout Piano/Assenze con righe giorni/bucket) ---
    wb = openpyxl.load_workbook(INPUT_XLSX)

    # --- Lettura tabelle con pandas ---
    buckets_df = pd.read_excel(INPUT_XLSX, sheet_name="Buckets")
    aule_df = pd.read_excel(INPUT_XLSX, sheet_name="Aule")
    edu_df = pd.read_excel(INPUT_XLSX, sheet_name="Educatori")
    tasks_df = pd.read_excel(INPUT_XLSX, sheet_name="Tasks")
    groups_df = pd.read_excel(INPUT_XLSX, sheet_name="GruppiU")
    vg_df = pd.read_excel(INPUT_XLSX, sheet_name="VincoliGenerali")

    days = ["Lunedi", "Martedi", "Mercoledi", "Giovedi", "Venerdi"]
    bucket_descr_to_type = dict(zip(buckets_df["Buck-Descr"], buckets_df["Buck-Type"]))

    # --- Struttura foglio Piano: mappa colonne e righe ---
    wsP = wb["Piano"]

    col_map = {}
    current_day = None
    for c in range(3, wsP.max_column + 1):
        d = wsP.cell(1, c).value
        if d is not None:
            current_day = str(d).strip()
        b = wsP.cell(2, c).value
        if b is None or current_day is None:
            continue
        b = str(b).strip()
        col_map[c] = (current_day, b, bucket_descr_to_type.get(b))

    row_map = {}
    r = 3
    while True:
        aid = wsP.cell(r, 1).value
        if aid is None:
            break
        row_map[r] = str(aid).strip()
        r += 1

    aula_ids = list(row_map.values())

    # slots = (day, bucket_descr, bucket_type, aula, piano_row, piano_col)
    slots = []
    for row, aula in row_map.items():
        for c, (day, buck, bt) in col_map.items():
            slots.append((day, buck, bt, aula, row, c))

    # --- Presenze/assenze ---
    pres_user = parse_matrix_presence(wb["AssenzeUtenti"])
    pres_edu = parse_matrix_presence(wb["AssenzeEducatori"])

    # --- Gruppi utenti ---
    group_users = {}
    for _, row in groups_df.iterrows():
        gid = str(row["Gruppo-ID"]).strip()
        users = set()
        for col in groups_df.columns:
            if str(col).startswith("UtID"):
                v = row[col]
                if pd.notna(v) and str(v).strip() != "":
                    users.add(str(v).strip())
        group_users[gid] = users

    # --- Indisponibilità aule / educatori (liste) ---
    aula_unavail_days = {str(r["Aula-ID"]).strip(): set(parse_list_cell(r.get("Giorni Indisponibili"))) for _, r in aule_df.iterrows()}
    aula_unavail_buckets = {str(r["Aula-ID"]).strip(): set(parse_list_cell(r.get("Buckets Indisponibili"))) for _, r in aule_df.iterrows()}

    edu_unavail_days = {str(r["Educatore-ID"]).strip(): set(parse_list_cell(r.get("Giorni Indisponibili"))) for _, r in edu_df.iterrows()}
    edu_unavail_buckets = {str(r["Educatore-ID"]).strip(): set(parse_list_cell(r.get("Buckets Indisponibili"))) for _, r in edu_df.iterrows()}

    edu_ids = [str(x).strip() for x in edu_df["Educatore-ID"].tolist()]

    vg1 = float(vg_df.loc[vg_df["Vinc-Gen-ID"] == "VG1", "Valore1"].iloc[0])
    vg2 = float(vg_df.loc[vg_df["Vinc-Gen-ID"] == "VG2", "Valore1"].iloc[0])

    # --- Espansione tasks in istanze (Molteplicità) ---
    instances = []
    for _, row in tasks_df.iterrows():
        tid = str(row["Task-ID"]).strip()
        mult = int(row["Molteplicità"])
        task_types = set(parse_int_list(row["Task-Type"]))
        bucket_pref = set(parse_int_list(row.get("Bucket-pref")))
        aul_pref = set(parse_id_list(row.get("Aule-Pref")))
        aul_alt = set(parse_id_list(row.get("Aula-Alt")))
        n_edu = int(row["Numero-Educatori"])
        edu_pref = set(parse_id_list(row.get("Educ-Pref")))
        edu_alt = set(parse_id_list(row.get("Educ-Alt")))
        edu_ris = set(parse_id_list(row.get("Educ-Ris")))
        group = str(row["Gruppo-ID"]).strip()
        min_users = int(row.get("#Min-Allievi", 0))

        for m in range(mult):
            instances.append(
                dict(
                    task_id=tid,
                    inst_id=f"{tid}__{m+1}",
                    task_types=task_types,
                    bucket_pref=bucket_pref,
                    aul_pref=aul_pref,
                    aul_alt=aul_alt,
                    n_edu=n_edu,
                    edu_pref=edu_pref,
                    edu_alt=edu_alt,
                    edu_ris=edu_ris,
                    group=group,
                    min_users=min_users,
                )
            )

    # --- Precompute: presenti per gruppo e timeslot (giorno,bucket_descr) ---
    # times = unici (day,bucket_descr,bucket_type) ricavati dalle colonne
    times = sorted(
        {(day, buck, bt) for (_, (day, buck, bt)) in col_map.items()},
        key=lambda x: (days.index(x[0]), x[2]),
    )

    present_count_group = defaultdict(dict)
    for gid, users in group_users.items():
        for day, buck, _ in times:
            cnt = sum(1 for u in users if pres_user.get(u, {}).get((day, buck), True))
            present_count_group[gid][(day, buck)] = cnt

    # --- Slot fattibili per istanza (vincoli bucket-type, aule pref/alt, indisponibilità, VG2) ---
    feasible_slots_per_inst = []
    pre_infeasible = []

    for inst in instances:
        feas = []
        allowed_aulas = inst["aul_pref"] | inst["aul_alt"]
        if not allowed_aulas:
            allowed_aulas = set(aula_ids)  # nessuna pref: tutte le aule

        for si, (day, buck, bt, aula, _, _) in enumerate(slots):
            if bt not in inst["task_types"]:
                continue
            if aula not in allowed_aulas:
                continue

            # aula indisponibile per giorno / bucket
            if day in aula_unavail_days.get(aula, set()):
                continue
            if buck in aula_unavail_buckets.get(aula, set()) or str(bt) in aula_unavail_buckets.get(aula, set()):
                continue

            # utenti effettivi (cap a min_allievi, ma se presenti meno di min -> usa presenti reali)
            users_avail = present_count_group[inst["group"]].get((day, buck), 0)
            if inst["min_users"] > 0:
                users_count = min(inst["min_users"], users_avail)
            else:
                users_count = users_avail

            # Vincolo VG2 (task): n_edu >= vg2 * users_count (altrimenti slot vietato)
            if users_count > 0 and inst["n_edu"] < vg2 * users_count - 1e-9:
                continue

            feas.append((si, users_count))

        feasible_slots_per_inst.append(feas)
        if not feas:
            pre_infeasible.append(inst["inst_id"])

    if pre_infeasible:
        print("Il piano non è fattibile: alcune istanze non hanno nessuno slot fattibile (vincoli bucket/aula/VG2).")
        print("Istanze non pianificabili:", pre_infeasible)
        return

    # --- Variabili x(ii,slot) e y(ii,slot,edu) ---
    x_entries = []
    for ii, inst in enumerate(instances):
        for si, users_count in feasible_slots_per_inst[ii]:
            day, buck, bt, aula, _, _ = slots[si]

            # costi x: bucket non preferito (2) se esiste Bucket-pref e bt non è nella lista
            bucket_cost = 0
            if inst["bucket_pref"] and bt not in inst["bucket_pref"]:
                bucket_cost = 2

            # costi x: aula alternativa (2) se esiste Aule-Pref e aula non è in pref
            aula_cost = 0
            if inst["aul_pref"] and aula not in inst["aul_pref"]:
                aula_cost = 2

            x_entries.append((ii, si, users_count, bucket_cost + aula_cost))

    x_index = {(ii, si): idx for idx, (ii, si, _, _) in enumerate(x_entries)}
    n_x = len(x_entries)

    # y: assegnazione educatore
    y_entries = []
    for xidx, (ii, si, _, _) in enumerate(x_entries):
        inst = instances[ii]
        day, buck, bt, _, _, _ = slots[si]

        if inst["edu_pref"]:
            allowed = inst["edu_pref"] | inst["edu_alt"] | inst["edu_ris"]
            if not allowed:
                allowed = set(edu_ids)
        else:
            allowed = set(edu_ids)

        for e in allowed:
            if e not in edu_ids:
                continue

            # indisponibilità educatore (anagrafica)
            if day in edu_unavail_days.get(e, set()):
                continue
            if buck in edu_unavail_buckets.get(e, set()) or str(bt) in edu_unavail_buckets.get(e, set()):
                continue

            # assenze educatore (matrice)
            if not pres_edu.get(e, {}).get((day, buck), True):
                continue

            # costo y: 1 sempre, +2 se pref esiste e e non è pref (alt/ris)
            base = 1
            extra = 0
            if inst["edu_pref"] and e not in inst["edu_pref"]:
                extra = 2

            y_entries.append((ii, si, e, xidx, base + extra))

    n_y = len(y_entries)
    n_vars = n_x + n_y

    # --- Obiettivo ---
    c = np.zeros(n_vars, dtype=float)
    for xidx, (_, _, _, cost) in enumerate(x_entries):
        c[xidx] = cost
    for yj, (_, _, _, _, cost) in enumerate(y_entries):
        c[n_x + yj] = cost

    # --- Vincoli ---
    A_data, A_row, A_col = [], [], []
    lb, ub = [], []
    row = 0

    # 1) ogni istanza assegnata esattamente una volta
    inst_to_xcols = defaultdict(list)
    for xidx, (ii, si, _, _) in enumerate(x_entries):
        inst_to_xcols[ii].append(xidx)

    for ii in range(len(instances)):
        cols = inst_to_xcols[ii]
        for col in cols:
            A_row.append(row); A_col.append(col); A_data.append(1.0)
        lb.append(1.0); ub.append(1.0)
        row += 1

    # 2) capacità slot (stesso giorno/bucket/aula): max 1 task
    slot_to_xcols = defaultdict(list)
    for xidx, (_, si, _, _) in enumerate(x_entries):
        slot_to_xcols[si].append(xidx)

    for si, cols in slot_to_xcols.items():
        for col in cols:
            A_row.append(row); A_col.append(col); A_data.append(1.0)
        lb.append(0.0); ub.append(1.0)
        row += 1

    # 3) numero educatori: sum y = n_edu * x per (ii,si)
    pair_to_y = defaultdict(list)
    for yj, (ii, si, _, _, _) in enumerate(y_entries):
        pair_to_y[(ii, si)].append(yj)

    for xidx, (ii, si, _, _) in enumerate(x_entries):
        yjs = pair_to_y.get((ii, si), [])
        for yj in yjs:
            A_row.append(row); A_col.append(n_x + yj); A_data.append(1.0)
        A_row.append(row); A_col.append(xidx); A_data.append(-float(instances[ii]["n_edu"]))
        lb.append(0.0); ub.append(0.0)
        row += 1

    # 4) y <= x
    for yj, (_, _, _, xidx, _) in enumerate(y_entries):
        A_row.append(row); A_col.append(n_x + yj); A_data.append(1.0)
        A_row.append(row); A_col.append(xidx); A_data.append(-1.0)
        lb.append(-np.inf); ub.append(0.0)
        row += 1

    # 5) educatore non usato due volte nello stesso (giorno,bucket) su qualunque aula
    # mappa xidx -> (day,buck)
    x_time = {}
    for xidx, (_, si, _, _) in enumerate(x_entries):
        day, buck, _, _, _, _ = slots[si]
        x_time[xidx] = (day, buck)

    edu_time_to_ycols = defaultdict(list)
    for yj, (_, _, e, xidx, _) in enumerate(y_entries):
        day, buck = x_time[xidx]
        edu_time_to_ycols[(e, day, buck)].append(n_x + yj)

    for _, cols in edu_time_to_ycols.items():
        for col in cols:
            A_row.append(row); A_col.append(col); A_data.append(1.0)
        lb.append(0.0); ub.append(1.0)
        row += 1

    # 6) vincolo gruppi (task-type 2/4): stesso utente non può essere in due tasks nello stesso giorno/bucket/aula
    special_types = {2, 4}
    inst_special = [(len(inst["task_types"] & special_types) > 0) for inst in instances]

    user_slot_to_xcols = defaultdict(list)
    for xidx, (ii, si, _, _) in enumerate(x_entries):
        if not inst_special[ii]:
            continue
        day, buck, _, aula, _, _ = slots[si]
        gid = instances[ii]["group"]
        for u in group_users.get(gid, set()):
            if pres_user.get(u, {}).get((day, buck), True):
                user_slot_to_xcols[(u, day, buck, aula)].append(xidx)

    for _, cols in user_slot_to_xcols.items():
        if len(cols) <= 1:
            continue
        for col in cols:
            A_row.append(row); A_col.append(col); A_data.append(1.0)
        lb.append(0.0); ub.append(1.0)
        row += 1

    # 7) VG1: totale educatori (bucket) / totale utenti (bucket) >= vg1  =>  sum_y - vg1*sum(users*x) >= 0
    time_to_x = defaultdict(list)   # (day,buck) -> list of (xidx, users_count)
    time_to_y = defaultdict(list)   # (day,buck) -> list of ycols
    for xidx, (_, _, users_count, _) in enumerate(x_entries):
        day, buck = x_time[xidx]
        time_to_x[(day, buck)].append((xidx, users_count))
    for yj, (_, _, _, xidx, _) in enumerate(y_entries):
        day, buck = x_time[xidx]
        time_to_y[(day, buck)].append(n_x + yj)

    for tkey in time_to_x.keys():
        for col in time_to_y.get(tkey, []):
            A_row.append(row); A_col.append(col); A_data.append(1.0)
        for xidx, users_count in time_to_x[tkey]:
            if users_count != 0:
                A_row.append(row); A_col.append(xidx); A_data.append(-vg1 * float(users_count))
        lb.append(0.0); ub.append(np.inf)
        row += 1

    A = sp.coo_matrix((A_data, (A_row, A_col)), shape=(row, n_vars)).tocsr()
    constraints = LinearConstraint(A, np.array(lb, dtype=float), np.array(ub, dtype=float))
    bounds = Bounds(np.zeros(n_vars), np.ones(n_vars))

    # tutte variabili binarie
    integrality = np.ones(n_vars, dtype=int)

    # --- Solve ---
    res = milp(c=c, constraints=constraints, integrality=integrality, bounds=bounds, options={"disp": False})

    if res.status != 0:
        # infeasible / altro
        if res.status == 2:
            print("Il piano non è fattibile (MILP infeasible).")
        else:
            print("Ottimizzazione non conclusa con ottimo. Status:", res.status)
        print(res.message)
        return

    x_sol = res.x[:n_x]
    y_sol = res.x[n_x:]

    # --- Ricostruzione piano ---
    # instanza -> slot scelto
    inst_to_slot = {}
    for xidx, val in enumerate(x_sol):
        if val > 0.5:
            ii, si, _, _ = x_entries[xidx]
            inst_to_slot[ii] = si

    # educatori per (ii,si)
    inst_slot_to_edus = defaultdict(list)
    for yj, val in enumerate(y_sol):
        if val > 0.5:
            ii, si, e, _, _ = y_entries[yj]
            inst_slot_to_edus[(ii, si)].append(e)

    # cell_text[(row,col)] = "Task-ID\nEd1, Ed2"
    cell_text = {}
    for ii, si in inst_to_slot.items():
        day, buck, _, aula, prow, pcol = slots[si]
        task_id = instances[ii]["task_id"]
        eds = sorted(inst_slot_to_edus[(ii, si)])
        cell_text[(prow, pcol)] = f"{task_id}\n{', '.join(eds)}"

    # --- Scrittura output excel ---
    out_wb = openpyxl.load_workbook(INPUT_XLSX)
    ws_out = out_wb["Piano"]

    wrap_align = Alignment(wrap_text=True, vertical="top")

    # pulizia e wrap su tutte le celle del piano
    for rr in range(3, ws_out.max_row + 1):
        if ws_out.cell(rr, 1).value is None:
            break
        for cc in range(3, ws_out.max_column + 1):
            cell = ws_out.cell(rr, cc)
            cell.value = None
            cell.alignment = wrap_align

    # scrivi valori
    for (rr, cc), txt in cell_text.items():
        cell = ws_out.cell(rr, cc)
        cell.value = txt
        cell.alignment = wrap_align

    # altezza righe (2 righe testo)
    for rr in range(3, ws_out.max_row + 1):
        if ws_out.cell(rr, 1).value is None:
            break
        ws_out.row_dimensions[rr].height = 30

    out_wb.save(OUTPUT_XLSX)
    print("Creato:", OUTPUT_XLSX)


if __name__ == "__main__":
    main()
